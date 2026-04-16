"""
KKC & Associates LLP — JM Expenses Analysis Tool
JMFSL / General Purpose
Streamlit app for Other Expenses audit testing:
  - Daybook upload & GL mapping from TB
  - Lead schedule, Variance analysis, GL detail sheets
  - Unusual items flagging
  - Auto-generated audit procedures & variance reasons
  - KKC-formatted Excel output
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ──────────────────────────────────────────────────────────────────

KKC_GREEN = "FF7CB542"
KKC_GREY = "FF808285"
WHITE = "FFFFFFFF"
YELLOW_FILL = "FFFFFF00"
LIGHT_GREEN_FILL = "FFE2EFDA"
LIGHT_GREY_FILL = "FFF2F2F2"

HEADER_FONT = Font(name="Source Sans Pro", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=KKC_GREEN, end_color=KKC_GREEN, fill_type="solid")
DATA_FONT = Font(name="Source Sans Pro", size=11)
BOLD_FONT = Font(name="Source Sans Pro", size=11, bold=True)
TITLE_FONT = Font(name="Source Sans Pro", size=14, bold=True, color=KKC_GREEN)
SUBTITLE_FONT = Font(name="Source Sans Pro", size=11, bold=True, color=KKC_GREY)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FMT = '#,##0'
PCT_FMT = '0.00%'

# Category display names mapped to grouping codes
CATEGORY_MAP = {
    "EOTHEXPRENT": "Rents & Related cost",
    "EOTHEXPR&T": "Rates & Taxes",
    "EOTHEXPR&M": "Repairs & Maint.",
    "EOTHEXPPROFFEES": "Legal & prof fees",
    "EOTHEXPINFOTECH": "IT Expenses",
    "EOTHEXPPRINTSTAT": "Printing & Stat.",
    "EOTHEXPEXTC": "Exchange Transaction",
    "EOTHEXPCOMMEXP": "Communication exp",
    "EOTHEXPMEMSUB": "Membership & subscription",
    "EOTHEXPELEC": "Electricity exp",
    "EOTHEXPINSU": "Insurance Exp",
    "EOTHEXPDONATION": "Donations",
    "EOTHEXPDIRCOMM": "Directors Commission",
    "EOTHEXPMANPOWER": "Manpower exp",
    "EOTHEXPTRAVEL": "Travelling & conveyance exp",
    "EOTHEXPBADDEBT": "Bad debts WO",
    "EOTHEXPADVERT": "Business development exp",
    "EOTHEXPSEVENT": "Expenditure on special events",
    "EOTHEXPRECRUITMENT": "Recruitment charges",
    "EOTHEXPMISCEXP": "Misc Exp",
    "EOTHEXPRECOVERY": "Recovery of expenses",
}

# Audit procedures per category
AUDIT_PROCEDURES = {
    "EOTHEXPRENT": [
        "Verify lease/rent agreements for all premises and check validity period.",
        "Confirm rent amounts match agreement terms; check for escalation clauses.",
        "Verify TDS deduction u/s 194-I on rent payments and timely deposit.",
        "Check Ind AS 116 lease classification and right-of-use asset accounting.",
        "Confirm security deposit balances and recoverability.",
        "Vouch significant payments to bank statements and landlord receipts.",
    ],
    "EOTHEXPR&T": [
        "Verify nature of rates and taxes — stamp duty, property tax, GST, professional tax, etc.",
        "Check whether amounts are in line with statutory rates applicable.",
        "Verify timely payment to avoid interest/penalty accruals.",
        "Confirm input tax credit eligibility where applicable.",
        "Vouch payments to challans and receipts from authorities.",
    ],
    "EOTHEXPR&M": [
        "Distinguish between revenue repairs and capital expenditure (Ind AS 16).",
        "Verify AMC contracts and match payments to contract terms.",
        "Check whether repairs relate to owned vs leased assets.",
        "Vouch significant payments to vendor invoices and work completion certificates.",
        "Verify TDS deduction u/s 194C on contractor payments.",
    ],
    "EOTHEXPPROFFEES": [
        "Obtain list of all professional/legal service providers engaged during the period.",
        "Verify fee agreements, engagement letters, and board approvals where required.",
        "Check TDS deduction u/s 194J on professional fees.",
        "Verify related party transactions for professional services (Ind AS 24).",
        "Vouch significant payments to invoices, fee notes, and bank statements.",
        "Check compliance with SEBI regulations for regulatory advisory fees.",
    ],
    "EOTHEXPINFOTECH": [
        "Obtain list of IT vendors and verify software license agreements.",
        "Distinguish between capital expenditure (intangible assets) and revenue expense.",
        "Verify AMC/SaaS subscription validity periods and expense recognition.",
        "Check TDS deduction on IT service payments.",
        "Verify data security and compliance-related IT expenditure.",
        "Vouch significant payments to vendor invoices and delivery confirmations.",
    ],
    "EOTHEXPPRINTSTAT": [
        "Verify nature of printing and stationery expenses.",
        "Check reasonableness by comparing with prior period trends.",
        "Vouch significant payments to vendor invoices.",
        "Verify whether any items should be capitalised (e.g., bulk IT consumables).",
    ],
    "EOTHEXPEXTC": [
        "Verify exchange gain/loss computation methodology.",
        "Check mark-to-market adjustments on outstanding foreign currency items (Ind AS 21).",
        "Reconcile to treasury/dealing desk records.",
        "Verify bank charges and forex conversion rates applied.",
        "Check for hedging transactions and hedge accounting compliance (Ind AS 109).",
    ],
    "EOTHEXPCOMMEXP": [
        "Verify telephone, internet, and courier service agreements.",
        "Check reasonableness of communication expenses month-on-month.",
        "Vouch significant payments to vendor bills and service provider statements.",
        "Verify TDS compliance on applicable communication services.",
    ],
    "EOTHEXPMEMSUB": [
        "Obtain list of memberships (exchange memberships, industry bodies, clubs).",
        "Verify membership fees to subscription confirmations and renewal notices.",
        "Check if any membership fees are capital in nature (exchange seats/cards).",
        "Verify board approval for significant memberships.",
    ],
    "EOTHEXPELEC": [
        "Verify electricity bills match metered consumption at premises.",
        "Check for any common area maintenance (CAM) charges included.",
        "Compare month-on-month consumption for reasonableness.",
        "Vouch payments to utility bills and bank statements.",
    ],
    "EOTHEXPINSU": [
        "Obtain schedule of all insurance policies (fire, D&O, professional indemnity, etc.).",
        "Verify premium amounts to policy documents and renewal notices.",
        "Check prepaid insurance computation for unexpired period (Ind AS compliant cut-off).",
        "Verify adequacy of insurance coverage for assets and operations.",
        "Confirm claims receivable, if any, and their recoverability.",
    ],
    "EOTHEXPDONATION": [
        "Verify board/CSR committee approval for all donations.",
        "Check whether donations qualify as CSR expenditure under Section 135.",
        "Verify 80G receipts from donee institutions where applicable.",
        "Check compliance with FCRA if donation to foreign entities.",
        "Vouch to bank statements and donation receipts.",
    ],
    "EOTHEXPDIRCOMM": [
        "Verify board resolution approving directors' commission/sitting fees.",
        "Check compliance with Companies Act 2013 Section 197 limits.",
        "Verify TDS deduction u/s 194J on directors' fees.",
        "Cross-check with related party disclosure requirements (Ind AS 24).",
        "Confirm amounts agree to Form MGT-7 / Board Report disclosures.",
    ],
    "EOTHEXPMANPOWER": [
        "Verify manpower outsourcing agreements and vendor contracts.",
        "Check TDS deduction u/s 194C on manpower supply payments.",
        "Verify headcount reconciliation between HR records and vendor bills.",
        "Check for employee vs contractor classification issues.",
        "Vouch significant payments to vendor invoices and attendance records.",
        "Verify PF/ESI compliance by manpower vendors where applicable.",
    ],
    "EOTHEXPTRAVEL": [
        "Verify travel policy compliance for claimed expenses.",
        "Check supporting documents — tickets, hotel bills, boarding passes.",
        "Verify approvals for foreign travel from competent authority.",
        "Check GST input credit on eligible travel expenses.",
        "Compare employee-wise and month-wise trends for reasonableness.",
        "Vouch significant claims to expense reports and bank reimbursements.",
    ],
    "EOTHEXPBADDEBT": [
        "Obtain ageing analysis of receivables written off.",
        "Verify board approval for write-off of bad debts.",
        "Check whether adequate provision (ECL) existed before write-off (Ind AS 109).",
        "Verify compliance with Section 36(1)(vii) for tax deductibility.",
        "Check if any amounts written off were subsequently recovered.",
        "Verify reversal of GST output tax on bad debts where applicable.",
    ],
    "EOTHEXPADVERT": [
        "Verify nature of business development expenses — advertising, client entertainment, events.",
        "Check board/management approvals for significant marketing campaigns.",
        "Verify TDS deduction on agency/vendor payments.",
        "Check whether any expenses should be classified as CSR.",
        "Vouch significant payments to vendor invoices and campaign reports.",
    ],
    "EOTHEXPSEVENT": [
        "Verify nature and purpose of special events.",
        "Check management/board approval for event expenditure.",
        "Verify vendor invoices and event completion reports.",
        "Check whether expenses relate to employee welfare or marketing.",
        "Vouch significant payments to bank statements.",
    ],
    "EOTHEXPRECRUITMENT": [
        "Verify recruitment agency agreements and fee structures.",
        "Check TDS deduction on recruitment consultancy fees.",
        "Verify reasonableness by comparing headcount additions to recruitment costs.",
        "Vouch significant payments to agency invoices and offer letters.",
    ],
    "EOTHEXPMISCEXP": [
        "Obtain breakup of miscellaneous expenses by nature.",
        "Verify that no items are misclassified under miscellaneous to avoid scrutiny.",
        "Check for items above materiality that should be separately disclosed.",
        "Vouch a sample of items to supporting invoices and approvals.",
        "Verify period-end accruals and their reversal in subsequent period.",
    ],
    "EOTHEXPRECOVERY": [
        "Verify nature of expense recoveries — inter-company, employee, vendor.",
        "Check credit notes / debit notes supporting the recovery.",
        "Verify GST treatment on recoveries (whether supply of service).",
        "Confirm amounts against underlying agreements or board approvals.",
        "Reconcile recoveries to related party transaction register (Ind AS 24).",
    ],
}

# ─── Helper Functions ───────────────────────────────────────────────────────────

def load_daybook(uploaded_files):
    """Load and concatenate all daybook Excel files (uses first sheet regardless of name)."""
    frames = []
    errors = []
    for f in uploaded_files:
        fname = f.name if hasattr(f, "name") else str(f)
        try:
            df = pd.read_excel(f, sheet_name=0, engine="openpyxl")
            frames.append(df)
        except Exception as e:
            errors.append(f"Could not read {fname}: {e}")
    if errors:
        import streamlit as _st
        for err in errors:
            _st.warning(err)
    if not frames:
        raise ValueError("No daybook files could be loaded successfully.")
    combined = pd.concat(frames, ignore_index=True)
    combined["Posting Date"] = pd.to_datetime(combined["Posting Date"], errors="coerce")
    return combined


def load_tb(uploaded_file):
    """Load TB sheet and Expenses sheet from TB file."""
    # Read all sheet names and match case-insensitively
    xl = pd.ExcelFile(uploaded_file, engine="openpyxl")
    sheet_map = {s.lower().strip(): s for s in xl.sheet_names}

    tb_key = sheet_map.get("tb")
    exp_key = sheet_map.get("expenses") or sheet_map.get("expense")

    if not tb_key:
        raise ValueError(
            f"Sheet 'TB' not found in uploaded file. "
            f"Available sheets: {xl.sheet_names}"
        )
    if not exp_key:
        raise ValueError(
            f"Sheet 'Expenses' not found in uploaded file. "
            f"Available sheets: {xl.sheet_names}"
        )

    # Auto-detect header row for TB sheet by searching for "GL code" / "New code"
    tb_raw = pd.read_excel(xl, sheet_name=tb_key, header=None, nrows=10)
    tb_header_row = None
    for i, row in tb_raw.iterrows():
        row_vals = [str(v).lower().strip() for v in row.values if pd.notna(v)]
        if any("gl code" in v for v in row_vals) and any("new code" in v for v in row_vals):
            tb_header_row = i
            break
    if tb_header_row is None:
        tb_header_row = 1  # fallback to original default

    tb = pd.read_excel(xl, sheet_name=tb_key, header=tb_header_row)

    # Auto-detect header row for Expenses sheet by searching for "Grouping Code" / "Particulars"
    exp_raw = pd.read_excel(xl, sheet_name=exp_key, header=None, nrows=10)
    exp_header_row = None
    for i, row in exp_raw.iterrows():
        row_vals = [str(v).lower().strip() for v in row.values if pd.notna(v)]
        if any("grouping" in v for v in row_vals) or any("particulars" in v for v in row_vals):
            exp_header_row = i
            break
    if exp_header_row is None:
        exp_header_row = 2  # fallback to original default

    expenses = pd.read_excel(xl, sheet_name=exp_key, header=exp_header_row)
    return tb, expenses


def detect_quarters(daybook):
    """Detect quarters present in daybook based on posting dates."""
    daybook = daybook.dropna(subset=["Posting Date"])
    daybook["YearMonth"] = daybook["Posting Date"].dt.to_period("M")

    # Define FY quarters (Apr-Mar)
    def get_fy_quarter(dt):
        m = dt.month
        if m in (4, 5, 6):
            return "Q1"
        elif m in (7, 8, 9):
            return "Q2"
        elif m in (10, 11, 12):
            return "Q3"
        elif m in (1, 2, 3):
            return "Q4"
        return None

    daybook["Quarter"] = daybook["Posting Date"].apply(get_fy_quarter)
    quarters_present = sorted(daybook["Quarter"].dropna().unique().tolist())
    return quarters_present


def get_quarter_end_label(q):
    """Return quarter-end date label for display."""
    mapping = {
        "Q1": "30th June",
        "Q2": "30th September",
        "Q3": "31st December",
        "Q4": "31st March",
    }
    return mapping.get(q, q)


def get_quarter_end_col_label(q, fy_year_end):
    """Return column header like 'As on 31st December 25'."""
    mapping = {
        "Q1": f"As on 30th June {fy_year_end - 1}",
        "Q2": f"As on 30th September {fy_year_end - 1}",
        "Q3": f"As on 31st December {fy_year_end - 1}",
        "Q4": f"As on 31st March {fy_year_end}",
    }
    return mapping.get(q, q)


def get_quarter_end_detail_label(q, fy_year_end):
    """Return detail sheet column header like 'December 31, 2025'."""
    mapping = {
        "Q1": f"June 30, {fy_year_end - 1}",
        "Q2": f"September 30, {fy_year_end - 1}",
        "Q3": f"December 31, {fy_year_end - 1}",
        "Q4": f"March 31, {fy_year_end}",
    }
    return mapping.get(q, q)


def _normalize_gl(val):
    """Normalize a GL code to a canonical string to avoid int/float/str mismatches."""
    if pd.isna(val):
        return None
    # Handle numpy int64, float64, python int, float, str uniformly
    try:
        # If it can be interpreted as a number with no decimal part, store as int string
        num = float(val)
        if num == int(num):
            return str(int(num))
        return str(val).strip()
    except (ValueError, TypeError):
        return str(val).strip()


def _find_col(df, candidates):
    """Find the first matching column name from candidates (case-insensitive, stripped)."""
    col_map = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        match = col_map.get(cand.lower().strip())
        if match:
            return match
    return None


def map_gl_to_category(tb):
    """Create GL code -> (New code / grouping code, GL Name) mapping from TB."""
    # Flexible column name matching
    new_code_col = _find_col(tb, ["New code", "New Code", "Newcode", "new code", "GroupingCode", "Grouping Code"])
    gl_code_col = _find_col(tb, ["GL code", "GL Code", "GLcode", "gl code", "GL code "])
    gl_name_col = _find_col(tb, ["GL Name", "GL name", "GLName", "gl name", "Description"])

    if not new_code_col or not gl_code_col:
        raise ValueError(
            f"Required columns not found in TB sheet. "
            f"Need 'New code' and 'GL code'. "
            f"Available columns: {list(tb.columns)}"
        )

    mapping = {}
    for _, row in tb.iterrows():
        new_code = row.get(new_code_col)
        gl_code = row.get(gl_code_col)
        gl_name = row.get(gl_name_col, "") if gl_name_col else ""
        if pd.notna(new_code) and pd.notna(gl_code):
            mapping[_normalize_gl(gl_code)] = {"grouping_code": str(new_code).strip(), "gl_name": gl_name}
    return mapping


def filter_other_expenses(daybook, gl_mapping, valid_grouping_codes):
    """Filter daybook to only Other Expenses GL codes."""
    other_exp_gls = {
        gl for gl, info in gl_mapping.items()
        if info["grouping_code"] in valid_grouping_codes
    }
    # Find GL account column in daybook (flexible name matching)
    gl_col = _find_col(daybook, ["G/L Account No.", "G/L Account No", "GL Account No.", "GL Account No", "GL code"])
    if not gl_col:
        raise ValueError(
            f"GL Account column not found in daybook. "
            f"Available columns: {list(daybook.columns)}"
        )
    # Normalize daybook GL codes to match mapping keys
    daybook = daybook.copy()
    daybook["_GL_Norm"] = daybook[gl_col].apply(_normalize_gl)
    filtered = daybook[daybook["_GL_Norm"].isin(other_exp_gls)].copy()
    filtered["Grouping Code"] = filtered["_GL_Norm"].map(
        lambda x: gl_mapping.get(x, {}).get("grouping_code", "UNKNOWN")
    )
    filtered["GL Name"] = filtered["_GL_Norm"].map(
        lambda x: gl_mapping.get(x, {}).get("gl_name", "")
    )
    filtered.drop(columns=["_GL_Norm"], inplace=True)
    return filtered


def compute_cumulative_by_quarter(filtered_daybook, quarters, gl_mapping, valid_grouping_codes):
    """
    Compute cumulative amounts per GL per quarter-end from daybook.
    Since daybook has Debit/Credit, net amount = Debit - Credit (for expenses, debit is increase).
    We compute cumulative (running total) at each quarter end.
    """
    def get_fy_quarter(dt):
        m = dt.month
        if m in (4, 5, 6):
            return "Q1"
        elif m in (7, 8, 9):
            return "Q2"
        elif m in (10, 11, 12):
            return "Q3"
        elif m in (1, 2, 3):
            return "Q4"
        return None

    df = filtered_daybook.copy()
    df["Quarter"] = df["Posting Date"].apply(get_fy_quarter)

    # Net amount per entry (use Amount column if available, else Debit - Credit)
    if "Amount" in df.columns:
        df["Net"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    else:
        df["Debit Amount"] = pd.to_numeric(df["Debit Amount"], errors="coerce").fillna(0)
        df["Credit Amount"] = pd.to_numeric(df["Credit Amount"], errors="coerce").fillna(0)
        df["Net"] = df["Debit Amount"] - df["Credit Amount"]

    # Quarter ordering
    q_order = ["Q1", "Q2", "Q3", "Q4"]
    q_present = [q for q in q_order if q in quarters]

    # Normalize GL codes for consistent grouping
    df["_GL_Norm"] = df["G/L Account No."].apply(_normalize_gl)

    # Aggregate per GL per quarter
    quarterly = df.groupby(["_GL_Norm", "Quarter"])["Net"].sum().reset_index()
    quarterly_pivot = quarterly.pivot_table(
        index="_GL_Norm", columns="Quarter", values="Net", fill_value=0
    )

    # Compute cumulative across quarters
    for i, q in enumerate(q_present):
        if q not in quarterly_pivot.columns:
            quarterly_pivot[q] = 0
        if i > 0:
            prev_q = q_present[i - 1]
            quarterly_pivot[q] = quarterly_pivot[q] + quarterly_pivot[prev_q]

    # Add GL metadata
    quarterly_pivot = quarterly_pivot.reset_index()
    quarterly_pivot.rename(columns={"_GL_Norm": "G/L Account No."}, inplace=True)
    quarterly_pivot["Grouping Code"] = quarterly_pivot["G/L Account No."].map(
        lambda x: gl_mapping.get(x, {}).get("grouping_code", "UNKNOWN")
    )
    quarterly_pivot["GL Name"] = quarterly_pivot["G/L Account No."].map(
        lambda x: gl_mapping.get(x, {}).get("gl_name", "")
    )

    return quarterly_pivot, q_present


def build_lead(quarterly_pivot, q_present, expenses_df, fy_year_end):
    """Build lead schedule: category-level totals per quarter-end."""
    lead_data = []
    for _, row in expenses_df.iterrows():
        gc = row.iloc[0]  # Grouping Code
        particulars = row.iloc[1]  # Particulars
        if pd.isna(gc):
            continue
        gc = str(gc).strip()
        cat_data = quarterly_pivot[quarterly_pivot["Grouping Code"] == gc]
        entry = {"Grouping Code": gc, "Particulars": particulars}
        for q in q_present:
            col_label = get_quarter_end_col_label(q, fy_year_end)
            entry[col_label] = cat_data[q].sum() if q in cat_data.columns else 0
        lead_data.append(entry)

    lead_df = pd.DataFrame(lead_data)
    return lead_df


def build_variance(lead_df, q_present, fy_year_end):
    """Build variance analysis: Q-o-Q incremental amounts and variance."""
    var_data = []
    col_labels = [get_quarter_end_col_label(q, fy_year_end) for q in q_present]

    for _, row in lead_df.iterrows():
        entry = {"Particulars": row["Particulars"]}
        # Cumulative amounts
        for cl in col_labels:
            entry[cl] = row.get(cl, 0)

        # Incremental quarters
        q_incremental = {}
        for i, q in enumerate(q_present):
            cl = col_labels[i]
            if i == 0:
                q_incremental[q] = entry[cl]  # First quarter = cumulative itself
            else:
                prev_cl = col_labels[i - 1]
                q_incremental[q] = entry[cl] - entry[prev_cl]
            entry[q] = q_incremental[q]

        # Variance between last two quarters
        if len(q_present) >= 2:
            last_q = q_present[-1]
            prev_q = q_present[-2]
            entry["Variance"] = q_incremental[last_q] - q_incremental[prev_q]
            total_last = sum(
                (q_incremental[last_q] - q_incremental[prev_q])
                for _ in [1]
            )
        else:
            entry["Variance"] = 0

        var_data.append(entry)

    var_df = pd.DataFrame(var_data)

    # Variance % (of total variance)
    total_var = var_df["Variance"].sum()
    var_df["Variance %"] = var_df["Variance"].apply(
        lambda x: x / total_var if total_var != 0 else 0
    )

    return var_df


def auto_generate_reasons(quarterly_pivot, q_present, grouping_code, fy_year_end):
    """Auto-generate variance reasons for a category based on GL-level movements."""
    cat_data = quarterly_pivot[quarterly_pivot["Grouping Code"] == grouping_code].copy()
    if len(q_present) < 2 or cat_data.empty:
        return "Insufficient data for variance analysis."

    last_q = q_present[-1]
    prev_q = q_present[-2]

    # Compute incremental for last two quarters
    if len(q_present) >= 3:
        prev_prev_q = q_present[-3]
        cat_data["Last_Q_Inc"] = cat_data.get(last_q, 0) - cat_data.get(prev_q, 0)
        cat_data["Prev_Q_Inc"] = cat_data.get(prev_q, 0) - cat_data.get(prev_prev_q, 0)
    elif len(q_present) == 2:
        cat_data["Last_Q_Inc"] = cat_data.get(last_q, 0) - cat_data.get(prev_q, 0)
        cat_data["Prev_Q_Inc"] = cat_data.get(prev_q, 0)  # First quarter = cumulative
    else:
        return "Only one quarter available."

    cat_data["GL_Variance"] = cat_data["Last_Q_Inc"] - cat_data["Prev_Q_Inc"]
    total_cat_var = cat_data["GL_Variance"].sum()

    if total_cat_var == 0:
        return "No material variance observed during the period."

    # Top movers (by absolute variance)
    cat_data["Abs_Var"] = cat_data["GL_Variance"].abs()
    top_movers = cat_data.nlargest(3, "Abs_Var")

    reasons = []
    for _, gl_row in top_movers.iterrows():
        gl_name = gl_row["GL Name"]
        gl_var = gl_row["GL_Variance"]
        if abs(gl_var) < 1:
            continue
        direction = "increase" if gl_var > 0 else "decrease"
        pct_of_total = abs(gl_var / total_cat_var * 100) if total_cat_var != 0 else 0
        reasons.append(
            f"{gl_name}: Rs. {abs(gl_var):,.0f} {direction} "
            f"({pct_of_total:.0f}% of total category variance)"
        )

    if not reasons:
        return "No material GL-level variance identified."

    return "Variance primarily driven by — " + "; ".join(reasons) + "."


def detect_unusual_items(filtered_daybook, gl_mapping):
    """Flag unusual items in the daybook for audit attention."""
    df = filtered_daybook.copy()
    df["Debit Amount"] = pd.to_numeric(df.get("Debit Amount", 0), errors="coerce").fillna(0)
    df["Credit Amount"] = pd.to_numeric(df.get("Credit Amount", 0), errors="coerce").fillna(0)
    df["Abs Amount"] = df[["Debit Amount", "Credit Amount"]].max(axis=1)

    flags = []

    # 1. Round amount entries (exact multiples of 100,000)
    round_mask = (df["Abs Amount"] >= 100000) & (df["Abs Amount"] % 100000 == 0)
    for idx in df[round_mask].index:
        flags.append({"Index": idx, "Flag": "Round amount (multiple of Rs. 1,00,000)"})

    # 2. Month-end / quarter-end entries (last 2 days of month)
    df["Day"] = df["Posting Date"].dt.day
    df["MonthEnd"] = df["Posting Date"].dt.days_in_month
    month_end_mask = (df["MonthEnd"] - df["Day"]) <= 1
    quarter_months = [6, 9, 12, 3]
    qe_mask = month_end_mask & df["Posting Date"].dt.month.isin(quarter_months)
    for idx in df[qe_mask].index:
        flags.append({"Index": idx, "Flag": "Quarter-end entry (last 2 days of quarter)"})

    # 3. High value entries (top 1% by absolute amount per category)
    for gc in df["Grouping Code"].unique():
        cat_df = df[df["Grouping Code"] == gc]
        if len(cat_df) < 10:
            threshold = cat_df["Abs Amount"].max() * 0.9
        else:
            threshold = cat_df["Abs Amount"].quantile(0.99)
        high_mask = cat_df["Abs Amount"] >= max(threshold, 1)
        for idx in cat_df[high_mask].index:
            flags.append({"Index": idx, "Flag": "High value entry within category"})

    # 4. Credit entries in expense accounts (reversals / unusual)
    credit_mask = df["Credit Amount"] > 0
    for idx in df[credit_mask].index:
        flags.append({"Index": idx, "Flag": "Credit entry in expense account (possible reversal)"})

    # 5. Entries with no narration
    no_narr_mask = df["Common Narration"].isna() | (df["Common Narration"].astype(str).str.strip() == "")
    no_narr_mask = no_narr_mask & (df["Narration"].isna() | (df["Narration"].astype(str).str.strip() == ""))
    for idx in df[no_narr_mask].index:
        flags.append({"Index": idx, "Flag": "No narration / description"})

    if not flags:
        return pd.DataFrame()

    flags_df = pd.DataFrame(flags)
    # Aggregate flags per entry
    agg_flags = flags_df.groupby("Index")["Flag"].apply(lambda x: " | ".join(sorted(set(x)))).reset_index()
    result = df.loc[agg_flags["Index"]].copy()
    result["Unusual Flags"] = agg_flags["Flag"].values
    result["Category"] = result["Grouping Code"].map(CATEGORY_MAP).fillna(result["Grouping Code"])

    # Sort by absolute amount descending
    result = result.sort_values("Abs Amount", ascending=False)

    cols = [
        "Posting Date", "Voucher No.", "G/L Account No.", "GL Name",
        "Category", "Debit Amount", "Credit Amount",
        "Common Narration", "Narration", "Unusual Flags"
    ]
    cols = [c for c in cols if c in result.columns]
    return result[cols].reset_index(drop=True)


# ─── Excel Export ───────────────────────────────────────────────────────────────

def apply_header_style(ws, row, max_col):
    """Apply KKC green header style to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_data_style(ws, row, max_col, bold=False):
    """Apply data style to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BOLD_FONT if bold else DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool) and cell.number_format != '@':
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")


def set_col_widths(ws, widths):
    """Set column widths from dict {col_letter: width}."""
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def write_kkc_title(ws, row, client_name, period, section_name):
    """Write KKC branded title rows."""
    ws.cell(row=row, column=1, value="KKC & Associates LLP").font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value=f"Client: {client_name}").font = SUBTITLE_FONT
    ws.cell(row=row + 1, column=3, value=f"Period: {period}").font = SUBTITLE_FONT
    ws.cell(row=row + 2, column=1, value=section_name).font = SUBTITLE_FONT
    return row + 3  # Return next available row (with one blank line = header row)



def generate_output_excel(
    lead_df, var_df, quarterly_pivot, q_present, filtered_daybook,
    unusual_df, expenses_df, gl_mapping, valid_grouping_codes,
    client_name, period, fy_year_end, prepared_by, reviewed_by
):
    """Generate the complete KKC-formatted Excel output."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Lead ──
    ws = wb.create_sheet("Lead")
    title_end = write_kkc_title(ws, 1, client_name, period, "Other Expenses — Lead Schedule")
    header_row = title_end + 1  # Row 5

    # Headers
    lead_cols = list(lead_df.columns)
    for ci, col_name in enumerate(lead_cols, 1):
        ws.cell(row=header_row, column=ci, value=col_name)
    apply_header_style(ws, header_row, len(lead_cols))

    # Data
    for ri, (_, row) in enumerate(lead_df.iterrows(), header_row + 1):
        for ci, col_name in enumerate(lead_cols, 1):
            val = row[col_name]
            cell = ws.cell(row=ri, column=ci, value=val)
            if col_name == "Grouping Code":
                cell.value = str(val)
                cell.number_format = '@'
        apply_data_style(ws, ri, len(lead_cols))

    # Total row
    total_row = header_row + len(lead_df) + 1
    ws.cell(row=total_row, column=1, value="")
    ws.cell(row=total_row, column=2, value="Total")
    for ci, col_name in enumerate(lead_cols[2:], 3):
        ws.cell(row=total_row, column=ci, value=lead_df[col_name].sum())
    apply_data_style(ws, total_row, len(lead_cols), bold=True)

    set_col_widths(ws, {"A": 23, "B": 40, "C": 18, "D": 18, "E": 18, "F": 18})

    # Footer
    footer_row = total_row + 2
    ws.cell(row=footer_row, column=1, value=f"Prepared by: {prepared_by}").font = DATA_FONT
    ws.cell(row=footer_row + 1, column=1, value=f"Reviewed by: {reviewed_by}").font = DATA_FONT
    ws.cell(
        row=footer_row + 2, column=1,
        value=f"Date: {datetime.now().strftime('%d %B %Y')}"
    ).font = DATA_FONT

    # ── Sheet 2: Variance ──
    ws = wb.create_sheet("Variance")
    title_end = write_kkc_title(ws, 1, client_name, period, "Other Expenses — Variance Analysis")
    header_row = title_end + 1

    # Build variance columns
    cumulative_cols = [get_quarter_end_col_label(q, fy_year_end) for q in q_present]
    q_inc_cols = q_present  # Q1, Q2, Q3, Q4
    var_cols = (
        ["Particulars"] + list(reversed(cumulative_cols)) +
        list(reversed(q_inc_cols)) +
        ["Variance", "Variance %", "Reasons", "Audit Procedure", "Team Remarks"]
    )

    for ci, col_name in enumerate(var_cols, 1):
        ws.cell(row=header_row, column=ci, value=col_name)
    apply_header_style(ws, header_row, len(var_cols))

    for ri, (_, row) in enumerate(var_df.iterrows(), header_row + 1):
        particulars = row["Particulars"]
        ci = 1
        ws.cell(row=ri, column=ci, value=particulars)
        ci += 1
        # Cumulative (reversed = latest first)
        for cl in reversed(cumulative_cols):
            ws.cell(row=ri, column=ci, value=row.get(cl, 0))
            ci += 1
        # Incremental quarters (reversed)
        for q in reversed(q_inc_cols):
            ws.cell(row=ri, column=ci, value=row.get(q, 0))
            ci += 1
        # Variance
        ws.cell(row=ri, column=ci, value=row.get("Variance", 0))
        ci += 1
        # Variance %
        cell = ws.cell(row=ri, column=ci, value=row.get("Variance %", 0))
        cell.number_format = PCT_FMT
        ci += 1

        # Auto-generated reasons
        gc = lead_df.iloc[ri - header_row - 1]["Grouping Code"] if (ri - header_row - 1) < len(lead_df) else None
        if gc:
            reason = auto_generate_reasons(quarterly_pivot, q_present, gc, fy_year_end)
            ws.cell(row=ri, column=ci, value=reason)
        ci += 1

        # Audit procedure summary
        if gc and gc in AUDIT_PROCEDURES:
            proc = "; ".join(AUDIT_PROCEDURES[gc][:3])  # Top 3 procedures
            ws.cell(row=ri, column=ci, value=proc)
        ci += 1

        # Team Remarks (blank)
        ws.cell(row=ri, column=ci, value="")
        apply_data_style(ws, ri, len(var_cols))

    # Total row
    total_row = header_row + len(var_df) + 1
    ws.cell(row=total_row, column=1, value="Total")
    ci = 2
    for cl in reversed(cumulative_cols):
        ws.cell(row=total_row, column=ci, value=var_df[cl].sum() if cl in var_df.columns else 0)
        ci += 1
    for q in reversed(q_inc_cols):
        ws.cell(row=total_row, column=ci, value=var_df[q].sum() if q in var_df.columns else 0)
        ci += 1
    ws.cell(row=total_row, column=ci, value=var_df["Variance"].sum())
    ci += 1
    ws.cell(row=total_row, column=ci, value=1.0)
    ws.cell(row=total_row, column=ci).number_format = PCT_FMT
    apply_data_style(ws, total_row, len(var_cols), bold=True)

    widths = {"A": 40}
    for i in range(2, len(var_cols) + 1):
        letter = get_column_letter(i)
        if i <= 1 + len(cumulative_cols) + len(q_inc_cols):
            widths[letter] = 17
        elif letter == get_column_letter(len(var_cols) - 2):
            widths[letter] = 72
        elif letter == get_column_letter(len(var_cols) - 1):
            widths[letter] = 50
        elif letter == get_column_letter(len(var_cols)):
            widths[letter] = 25
        else:
            widths[letter] = 17
    set_col_widths(ws, widths)

    # ── Sheets 3+: Category Detail Sheets ──
    for _, exp_row in expenses_df.iterrows():
        gc = exp_row.iloc[0]
        if pd.isna(gc):
            continue
        gc = str(gc).strip()
        cat_name = CATEGORY_MAP.get(gc, gc)
        # Sheet name max 31 chars
        sheet_name = cat_name[:31]

        ws = wb.create_sheet(sheet_name)
        title_end = write_kkc_title(ws, 1, client_name, period, f"Other Expenses — {cat_name}")
        header_row = title_end + 1

        # GL detail for this category
        cat_data = quarterly_pivot[quarterly_pivot["Grouping Code"] == gc].copy()

        # Detail columns
        detail_cumul_labels = [get_quarter_end_detail_label(q, fy_year_end) for q in reversed(q_present)]
        detail_q_labels = [f"{q} (Incremental)" for q in reversed(q_present)]

        detail_headers = ["Grouping Code", "GL Code", "GL Name"] + detail_cumul_labels

        # Add incremental columns
        if len(q_present) >= 2:
            for q in reversed(q_present):
                detail_headers.append(q)
            detail_headers += ["Variance", "Reasons", "Team Remarks"]

        for ci, h in enumerate(detail_headers, 1):
            ws.cell(row=header_row, column=ci, value=h)
        apply_header_style(ws, header_row, len(detail_headers))

        ri = header_row + 1
        for _, gl_row in cat_data.iterrows():
            ci = 1
            cell = ws.cell(row=ri, column=ci, value=str(gc))
            cell.number_format = '@'
            ci += 1
            gl_code_val = gl_row["G/L Account No."]
            gl_code_str = str(int(gl_code_val)) if isinstance(gl_code_val, float) and gl_code_val == int(gl_code_val) else str(gl_code_val)
            cell = ws.cell(row=ri, column=ci, value=gl_code_str)
            cell.number_format = '@'
            ci += 1
            ws.cell(row=ri, column=ci, value=gl_row["GL Name"])
            ci += 1

            # Cumulative amounts (reversed = latest first)
            for q in reversed(q_present):
                ws.cell(row=ri, column=ci, value=gl_row.get(q, 0))
                ci += 1

            # Incremental quarters
            if len(q_present) >= 2:
                incremental_vals = {}
                for i, q in enumerate(q_present):
                    if i == 0:
                        incremental_vals[q] = gl_row.get(q, 0)
                    else:
                        prev_q = q_present[i - 1]
                        incremental_vals[q] = gl_row.get(q, 0) - gl_row.get(prev_q, 0)

                for q in reversed(q_present):
                    ws.cell(row=ri, column=ci, value=incremental_vals[q])
                    ci += 1

                # Variance (last Q incremental - prev Q incremental)
                last_q = q_present[-1]
                prev_q = q_present[-2]
                variance = incremental_vals[last_q] - incremental_vals[prev_q]
                ws.cell(row=ri, column=ci, value=variance)
                ci += 1

                # Reasons (blank for GL level - filled at category level in variance sheet)
                ws.cell(row=ri, column=ci, value="")
                ci += 1

                # Team Remarks
                ws.cell(row=ri, column=ci, value="")
                ci += 1

            apply_data_style(ws, ri, len(detail_headers))
            ri += 1

        # Total row for detail sheet
        if not cat_data.empty:
            ws.cell(row=ri, column=2, value="Total")
            ci = 4
            for q in reversed(q_present):
                ws.cell(row=ri, column=ci, value=cat_data[q].sum() if q in cat_data.columns else 0)
                ci += 1
            if len(q_present) >= 2:
                for q in reversed(q_present):
                    inc_sum = 0
                    for _, gl_row in cat_data.iterrows():
                        q_idx = q_present.index(q)
                        if q_idx == 0:
                            inc_sum += gl_row.get(q, 0)
                        else:
                            inc_sum += gl_row.get(q, 0) - gl_row.get(q_present[q_idx - 1], 0)
                    ws.cell(row=ri, column=ci, value=inc_sum)
                    ci += 1
                # Total variance
                last_q = q_present[-1]
                prev_q = q_present[-2]
                total_var = 0
                for _, gl_row in cat_data.iterrows():
                    li = q_present.index(last_q)
                    pi = q_present.index(prev_q)
                    last_inc = gl_row.get(last_q, 0) - (gl_row.get(q_present[li - 1], 0) if li > 0 else 0)
                    prev_inc = gl_row.get(prev_q, 0) - (gl_row.get(q_present[pi - 1], 0) if pi > 0 else 0)
                    total_var += last_inc - prev_inc
                ws.cell(row=ri, column=ci, value=total_var)
            apply_data_style(ws, ri, len(detail_headers), bold=True)
            ri += 1

        # Audit procedures section
        ri += 1
        ws.cell(row=ri, column=1, value="Audit Procedures").font = BOLD_FONT
        ws.cell(row=ri, column=1).fill = PatternFill(
            start_color=LIGHT_GREEN_FILL[2:], end_color=LIGHT_GREEN_FILL[2:], fill_type="solid"
        )
        ri += 1

        procedures = AUDIT_PROCEDURES.get(gc, ["Vouch a sample of transactions to supporting documents."])
        for pi, proc in enumerate(procedures, 1):
            ws.cell(row=ri, column=1, value=f"{pi}.")
            ws.cell(row=ri, column=2, value=proc).font = DATA_FONT
            ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=6)
            ri += 1

        # Variance-driven additional procedures
        if len(q_present) >= 2:
            total_cat_var = 0
            for _, gl_row in cat_data.iterrows():
                li = q_present.index(q_present[-1])
                pi = q_present.index(q_present[-2])
                last_inc = gl_row.get(q_present[-1], 0) - (gl_row.get(q_present[li - 1], 0) if li > 0 else 0)
                prev_inc = gl_row.get(q_present[-2], 0) - (gl_row.get(q_present[pi - 1], 0) if pi > 0 else 0)
                total_cat_var += last_inc - prev_inc

            if abs(total_cat_var) > 0:
                ri += 1
                ws.cell(row=ri, column=1, value="Variance-Driven Additional Procedures").font = BOLD_FONT
                ws.cell(row=ri, column=1).fill = PatternFill(
                    start_color=YELLOW_FILL[2:], end_color=YELLOW_FILL[2:], fill_type="solid"
                )
                ri += 1
                direction = "increase" if total_cat_var > 0 else "decrease"
                ws.cell(
                    row=ri, column=2,
                    value=f"Category shows net {direction} of Rs. {abs(total_cat_var):,.0f} Q-o-Q. "
                          f"Investigate top contributing GL accounts and obtain explanations from management."
                ).font = DATA_FONT
                ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=8)
                ri += 1

                reason_text = auto_generate_reasons(quarterly_pivot, q_present, gc, fy_year_end)
                ws.cell(row=ri, column=2, value=reason_text).font = DATA_FONT
                ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=8)

        # ── Top 3 Variance GL Pivot by OppGlName (Q-o-Q) ──
        if len(q_present) >= 2 and not cat_data.empty:
            # Compute variance per GL in this category
            gl_variances = []
            for _, gl_row in cat_data.iterrows():
                gl_code = gl_row["G/L Account No."]
                gl_name = gl_row["GL Name"]
                inc_vals = {}
                for i, q in enumerate(q_present):
                    if i == 0:
                        inc_vals[q] = gl_row.get(q, 0)
                    else:
                        inc_vals[q] = gl_row.get(q, 0) - gl_row.get(q_present[i - 1], 0)
                var = inc_vals[q_present[-1]] - inc_vals[q_present[-2]]
                gl_variances.append({"gl_code": gl_code, "gl_name": gl_name, "variance": var, "abs_var": abs(var)})

            gl_variances.sort(key=lambda x: x["abs_var"], reverse=True)
            top3 = gl_variances[:3]
            top3 = [g for g in top3 if g["abs_var"] > 0]

            if top3:
                ri += 2
                ws.cell(row=ri, column=1, value="Top Variance GL — OppGlName Pivot (Q-o-Q)").font = BOLD_FONT
                ws.cell(row=ri, column=1).fill = PatternFill(
                    start_color=LIGHT_GREEN_FILL[2:], end_color=LIGHT_GREEN_FILL[2:], fill_type="solid"
                )
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
                ri += 1

                # Prepare quarter labels for the pivot
                def _get_fy_quarter(dt):
                    m = dt.month
                    if m in (4, 5, 6): return "Q1"
                    elif m in (7, 8, 9): return "Q2"
                    elif m in (10, 11, 12): return "Q3"
                    elif m in (1, 2, 3): return "Q4"
                    return None

                for gl_info in top3:
                    gl_code = gl_info["gl_code"]
                    gl_name = gl_info["gl_name"]
                    gl_code_str = str(int(gl_code)) if isinstance(gl_code, float) and gl_code == int(gl_code) else str(gl_code)
                    direction = "increase" if gl_info["variance"] > 0 else "decrease"

                    # Sub-header for this GL
                    ri += 1
                    ws.cell(row=ri, column=1, value=f"GL: {gl_code_str} — {gl_name}").font = BOLD_FONT
                    ws.cell(row=ri, column=4, value=f"Variance: Rs. {abs(gl_info['variance']):,.0f} ({direction})").font = SUBTITLE_FONT
                    ri += 1

                    # Pivot headers: Sr. No. | OppGlName | Q1 | Q2 | ... | Variance (last Q - prev Q)
                    pivot_hdrs = ["Sr. No.", "OppGlName"]
                    for q in q_present:
                        pivot_hdrs.append(q)
                    pivot_hdrs.append(f"Variance ({q_present[-1]} - {q_present[-2]})")
                    num_pivot_cols = len(pivot_hdrs)

                    for ci, h in enumerate(pivot_hdrs, 1):
                        ws.cell(row=ri, column=ci, value=h)
                    apply_header_style(ws, ri, num_pivot_cols)
                    ri += 1

                    # Filter daybook for this GL and assign quarter
                    gl_daybook = filtered_daybook[filtered_daybook["G/L Account No."] == gl_code].copy()
                    if "OppGlName" in gl_daybook.columns and not gl_daybook.empty:
                        gl_daybook["Amount"] = pd.to_numeric(gl_daybook.get("Amount", 0), errors="coerce").fillna(0)
                        gl_daybook["_Q"] = gl_daybook["Posting Date"].apply(_get_fy_quarter)

                        # Pivot: OppGlName × Quarter
                        opp_q_pivot = gl_daybook.groupby(["OppGlName", "_Q"], dropna=False)["Amount"].sum().reset_index()
                        opp_q_wide = opp_q_pivot.pivot_table(
                            index="OppGlName", columns="_Q", values="Amount", fill_value=0, aggfunc="sum"
                        )
                        # Ensure all quarters present
                        for q in q_present:
                            if q not in opp_q_wide.columns:
                                opp_q_wide[q] = 0
                        opp_q_wide = opp_q_wide[q_present]

                        # Compute variance column (last Q - prev Q)
                        opp_q_wide["_Variance"] = opp_q_wide[q_present[-1]] - opp_q_wide[q_present[-2]]

                        # Sort by absolute variance descending
                        opp_q_wide = opp_q_wide.reindex(
                            opp_q_wide["_Variance"].abs().sort_values(ascending=False).index
                        )

                        for sr, (opp_name, opp_row) in enumerate(opp_q_wide.iterrows(), 1):
                            disp_name = str(opp_name) if pd.notna(opp_name) else "(Blank)"
                            ci = 1
                            ws.cell(row=ri, column=ci, value=sr); ci += 1
                            ws.cell(row=ri, column=ci, value=disp_name); ci += 1
                            for q in q_present:
                                ws.cell(row=ri, column=ci, value=opp_row[q]); ci += 1
                            ws.cell(row=ri, column=ci, value=opp_row["_Variance"])
                            apply_data_style(ws, ri, num_pivot_cols)
                            ri += 1

                        # Total row
                        ci = 1
                        ws.cell(row=ri, column=1, value=""); ci = 2
                        ws.cell(row=ri, column=ci, value="Total"); ci += 1
                        for q in q_present:
                            ws.cell(row=ri, column=ci, value=opp_q_wide[q].sum()); ci += 1
                        ws.cell(row=ri, column=ci, value=opp_q_wide["_Variance"].sum())
                        apply_data_style(ws, ri, num_pivot_cols, bold=True)
                        ri += 1
                    else:
                        ws.cell(row=ri, column=1, value="No daybook data available for this GL.").font = DATA_FONT
                        ri += 1

        # Column widths for detail sheets
        set_col_widths(ws, {
            "A": 20, "B": 45, "C": 35, "D": 18, "E": 18, "F": 18,
            "G": 18, "H": 18, "I": 18, "J": 50, "K": 25, "L": 25
        })

    # ── Unusual Items Sheet ──
    if not unusual_df.empty:
        ws = wb.create_sheet("Unusual Items")
        title_end = write_kkc_title(ws, 1, client_name, period, "Other Expenses — Unusual Items for Audit Attention")
        header_row = title_end + 1

        unusual_cols = list(unusual_df.columns)
        for ci, h in enumerate(unusual_cols, 1):
            ws.cell(row=header_row, column=ci, value=h)
        apply_header_style(ws, header_row, len(unusual_cols))

        max_unusual = min(len(unusual_df), 2000)  # Cap at 2000 rows
        for ri, (_, row) in enumerate(unusual_df.head(max_unusual).iterrows(), header_row + 1):
            for ci, col_name in enumerate(unusual_cols, 1):
                val = row[col_name]
                if isinstance(val, pd.Timestamp):
                    val = val.strftime("%d %B %Y")
                cell = ws.cell(row=ri, column=ci, value=val)
                if col_name == "G/L Account No.":
                    gl_str = str(int(val)) if isinstance(val, float) and val == int(val) else str(val)
                    cell.value = gl_str
                    cell.number_format = '@'
            apply_data_style(ws, ri, len(unusual_cols))

        set_col_widths(ws, {
            "A": 16, "B": 18, "C": 16, "D": 35, "E": 28,
            "F": 18, "G": 18, "H": 40, "I": 40, "J": 55
        })

        # Summary counts by flag type
        ri = header_row + max_unusual + 2
        ws.cell(row=ri, column=1, value="Summary of Unusual Items").font = BOLD_FONT
        ri += 1
        ws.cell(row=ri, column=1, value="Total unusual items flagged:").font = DATA_FONT
        ws.cell(row=ri, column=2, value=len(unusual_df)).font = BOLD_FONT

    # ── Audit Procedures Summary Sheet ──
    ws = wb.create_sheet("Audit Procedures")
    title_end = write_kkc_title(ws, 1, client_name, period, "Other Expenses — Audit Procedures Guide")
    header_row = title_end + 1

    ws.cell(row=header_row, column=1, value="Expense Category")
    ws.cell(row=header_row, column=2, value="Sr. No.")
    ws.cell(row=header_row, column=3, value="Audit Procedure")
    ws.cell(row=header_row, column=4, value="Completed (Y/N)")
    ws.cell(row=header_row, column=5, value="WP Ref")
    ws.cell(row=header_row, column=6, value="Team Remarks")
    apply_header_style(ws, header_row, 6)

    ri = header_row + 1
    for gc, procs in AUDIT_PROCEDURES.items():
        cat_name = CATEGORY_MAP.get(gc, gc)
        for pi, proc in enumerate(procs, 1):
            ws.cell(row=ri, column=1, value=cat_name if pi == 1 else "")
            ws.cell(row=ri, column=2, value=pi)
            ws.cell(row=ri, column=3, value=proc)
            ws.cell(row=ri, column=4, value="")  # Completed
            ws.cell(row=ri, column=5, value="")  # WP Ref
            ws.cell(row=ri, column=6, value="")  # Remarks
            apply_data_style(ws, ri, 6)
            if pi == 1:
                ws.cell(row=ri, column=1).font = BOLD_FONT
            ri += 1
        ri += 1  # Blank row between categories

    set_col_widths(ws, {"A": 32, "B": 10, "C": 80, "D": 16, "E": 12, "F": 30})

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── Streamlit App ──────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="KKC — JM Expenses Analysis Tool",
        page_icon="",
        layout="wide",
    )

    st.markdown(
        """
        <style>
        .main-header {
            font-family: 'Source Sans Pro', sans-serif;
            color: #7CB542;
            font-size: 28px;
            font-weight: bold;
        }
        .sub-header {
            font-family: 'Source Sans Pro', sans-serif;
            color: #808285;
            font-size: 16px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<p class="main-header">KKC & Associates LLP</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="sub-header">JM Expenses Analysis Tool — Vouching & Variance Analysis</p>',
        unsafe_allow_html=True,
    )
    st.markdown("---")

    # ── Sidebar: Engagement Details ──
    with st.sidebar:
        st.header("Engagement Details")
        client_name = st.text_input("Client Name", value="")
        period = st.text_input("Audit Period", value="FY 2025-26")
        fy_year_end = st.number_input("FY Year End (e.g., 2026 for March 2026)", value=2026, min_value=2020, max_value=2035)
        prepared_by = st.text_input("Prepared by", value="")
        reviewed_by = st.text_input("Reviewed by", value="")

        st.markdown("---")
        st.header("Upload Files")
        daybook_files = st.file_uploader(
            "Daybook Files (Excel)", type=["xlsx"], accept_multiple_files=True,
            help="Upload quarterly daybook files. Each file should have a 'data' sheet."
        )
        tb_file = st.file_uploader(
            "TB & Other Expenses Lead (Excel)", type=["xlsx"],
            help="Upload the TB file with 'TB' and 'Expenses' sheets."
        )

    # ── Main Area ──
    if not daybook_files or not tb_file:
        st.info("Upload daybook files and TB file from the sidebar to begin.")
        st.markdown("### How to use this tool")
        st.markdown("""
        1. Enter engagement details in the sidebar
        2. Upload one or more quarterly daybook Excel files
        3. Upload the TB & Other Expenses Lead file
        4. The tool will automatically:
           - Map GL codes to expense categories from the TB
           - Compute quarterly cumulative and incremental amounts
           - Perform Q-o-Q variance analysis with auto-generated reasons
           - Flag unusual items for audit attention
           - Generate audit procedures per expense category
        5. Download the KKC-formatted Excel output
        """)
        return

    # ── Processing ──
    with st.spinner("Loading daybook files..."):
        daybook = load_daybook(daybook_files)
        st.success(f"Daybook loaded: {len(daybook):,} entries from {len(daybook_files)} file(s)")

    with st.spinner("Loading TB and expense classification..."):
        tb, expenses_df = load_tb(tb_file)
        # Clean expenses_df
        expenses_df = expenses_df.dropna(subset=[expenses_df.columns[0]])
        # Remove total rows
        expenses_df = expenses_df[
            ~expenses_df.iloc[:, 1].astype(str).str.lower().str.contains("total", na=False)
        ]
        st.success(f"TB loaded: {len(tb):,} GL accounts | {len(expenses_df)} expense categories")

    # GL mapping
    gl_mapping = map_gl_to_category(tb)
    valid_grouping_codes = set(expenses_df.iloc[:, 0].dropna().astype(str).str.strip().tolist())

    # Detect quarters
    quarters = detect_quarters(daybook)
    st.info(f"Quarters detected in daybook: {', '.join(quarters)}")

    # ── Diagnostic: GL matching debug ──
    other_exp_gls = {gl for gl, info in gl_mapping.items() if info["grouping_code"] in valid_grouping_codes}
    db_gl_sample = daybook["G/L Account No."].dropna().head(3).tolist()
    db_gl_norm_sample = [_normalize_gl(x) for x in db_gl_sample]
    mapping_sample = list(other_exp_gls)[:3]
    with st.expander("Debug: GL Code Matching Info"):
        st.write(f"**TB GL code column dtype:** `{tb['GL code'].dtype}`")
        st.write(f"**Daybook G/L Account No. dtype:** `{daybook['G/L Account No.'].dtype}`")
        st.write(f"**Daybook GL samples (raw):** `{db_gl_sample}` → types: `{[type(x).__name__ for x in db_gl_sample]}`")
        st.write(f"**Daybook GL samples (normalized):** `{db_gl_norm_sample}`")
        st.write(f"**Mapping GL samples (normalized):** `{mapping_sample}` → types: `{[type(x).__name__ for x in mapping_sample]}`")
        st.write(f"**Other expense GL codes in mapping:** `{len(other_exp_gls)}`")
        st.write(f"**Valid grouping codes:** `{len(valid_grouping_codes)}`")
        # Check overlap
        db_norm_set = set(daybook["G/L Account No."].dropna().apply(_normalize_gl).unique())
        overlap = other_exp_gls & db_norm_set
        st.write(f"**Overlap (normalized):** `{len(overlap)}` GL codes match")
        if len(overlap) == 0:
            st.error("NO GL codes match between TB and daybook after normalization!")
            st.write(f"**TB 'GL code' column name check:** `{'GL code' in tb.columns}`")
            st.write(f"**TB columns:** `{list(tb.columns)[:10]}`")
            st.write(f"**Daybook columns:** `{list(daybook.columns)[:10]}`")

    # Filter other expenses
    with st.spinner("Filtering Other Expenses entries from daybook..."):
        filtered = filter_other_expenses(daybook, gl_mapping, valid_grouping_codes)
        # Normalize GL codes in filtered daybook to match quarterly_pivot keys
        filtered["G/L Account No."] = filtered["G/L Account No."].apply(_normalize_gl)
        st.success(f"Other Expenses entries: {len(filtered):,} out of {len(daybook):,} total")

    # Compute cumulative
    with st.spinner("Computing quarterly amounts..."):
        quarterly_pivot, q_present = compute_cumulative_by_quarter(
            filtered, quarters, gl_mapping, valid_grouping_codes
        )

    # Build lead
    lead_df = build_lead(quarterly_pivot, q_present, expenses_df, fy_year_end)

    # Build variance
    var_df = build_variance(lead_df, q_present, fy_year_end)

    # Detect unusual items
    with st.spinner("Flagging unusual items..."):
        unusual_df = detect_unusual_items(filtered, gl_mapping)

    # ── Display Tabs ──
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Lead Schedule", "Variance Analysis", "Category Details",
        "Unusual Items", "Audit Procedures"
    ])

    with tab1:
        st.subheader("Lead Schedule — Other Expenses")
        display_lead = lead_df.copy()
        # Add total row
        total_row = {"Grouping Code": "", "Particulars": "Total"}
        for col in display_lead.columns[2:]:
            total_row[col] = display_lead[col].sum()
        display_lead = pd.concat([display_lead, pd.DataFrame([total_row])], ignore_index=True)
        st.dataframe(display_lead, use_container_width=True, hide_index=True)

    with tab2:
        st.subheader("Variance Analysis — Q-o-Q")
        display_var = var_df.copy()
        # Add reasons column for display
        reasons_list = []
        for idx, row in lead_df.iterrows():
            gc = row["Grouping Code"]
            reason = auto_generate_reasons(quarterly_pivot, q_present, gc, fy_year_end)
            reasons_list.append(reason)
        display_var["Reasons"] = reasons_list[:len(display_var)]
        st.dataframe(display_var, use_container_width=True, hide_index=True)

    with tab3:
        st.subheader("Category-wise GL Detail")
        selected_cat = st.selectbox(
            "Select Category",
            options=expenses_df.iloc[:, 0].dropna().tolist(),
            format_func=lambda x: f"{x} — {CATEGORY_MAP.get(str(x).strip(), x)}"
        )
        if selected_cat:
            gc = str(selected_cat).strip()
            cat_detail = quarterly_pivot[quarterly_pivot["Grouping Code"] == gc].copy()

            # Compute incremental columns
            for i, q in enumerate(q_present):
                if i == 0:
                    cat_detail[f"{q}_inc"] = cat_detail[q]
                else:
                    cat_detail[f"{q}_inc"] = cat_detail[q] - cat_detail[q_present[i - 1]]

            display_cols = ["G/L Account No.", "GL Name"]
            for q in reversed(q_present):
                display_cols.append(q)
            if len(q_present) >= 2:
                for q in reversed(q_present):
                    display_cols.append(f"{q}_inc")

            st.dataframe(
                cat_detail[display_cols].reset_index(drop=True),
                use_container_width=True,
                hide_index=True,
            )

    with tab4:
        st.subheader("Unusual Items — Flagged for Audit Attention")
        if unusual_df.empty:
            st.success("No unusual items detected.")
        else:
            st.warning(f"{len(unusual_df):,} unusual items flagged")

            # Summary by flag type
            if "Unusual Flags" in unusual_df.columns:
                all_flags = unusual_df["Unusual Flags"].str.split(" \\| ").explode()
                flag_counts = all_flags.value_counts()
                st.markdown("**Flag Summary:**")
                for flag, count in flag_counts.items():
                    st.markdown(f"- {flag}: **{count:,}** items")

            st.dataframe(unusual_df.head(500), use_container_width=True, hide_index=True)
            if len(unusual_df) > 500:
                st.caption(f"Showing first 500 of {len(unusual_df):,} items. Full list available in Excel output.")

    with tab5:
        st.subheader("Audit Procedures — By Expense Category")
        for gc, procs in AUDIT_PROCEDURES.items():
            if gc in valid_grouping_codes:
                cat_name = CATEGORY_MAP.get(gc, gc)
                with st.expander(f"{cat_name}"):
                    for i, proc in enumerate(procs, 1):
                        st.markdown(f"{i}. {proc}")

    # ── Generate & Download ──
    st.markdown("---")
    st.subheader("Generate KKC Output")

    if st.button("Generate Excel Output", type="primary", use_container_width=True):
        with st.spinner("Generating KKC-formatted Excel output..."):
            output = generate_output_excel(
                lead_df=lead_df,
                var_df=var_df,
                quarterly_pivot=quarterly_pivot,
                q_present=q_present,
                filtered_daybook=filtered,
                unusual_df=unusual_df,
                expenses_df=expenses_df,
                gl_mapping=gl_mapping,
                valid_grouping_codes=valid_grouping_codes,
                client_name=client_name,
                period=period,
                fy_year_end=fy_year_end,
                prepared_by=prepared_by,
                reviewed_by=reviewed_by,
            )
        st.success("Output generated successfully!")
        st.download_button(
            label="Download KKC Output Excel",
            data=output,
            file_name=f"Other Expenses_KKC_output_{client_name.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.officedocument",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
